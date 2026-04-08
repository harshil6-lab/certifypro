"""Certificate generation service.

Downloads a template image from a URL, overlays student name / QR code /
certificate ID at positions derived from layout_config (percent values), and
writes the finished PNG to the local uploads/generated/ directory.

Dependencies (add to requirements.txt if not present):
    pillow, qrcode[pil], openpyxl, httpx
"""

from __future__ import annotations

import io
import os
import uuid
from typing import Any

import httpx
import openpyxl
import qrcode
from PIL import Image, ImageDraw, ImageFont

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_BACKEND_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
GENERATED_DIR = os.path.join(_BACKEND_ROOT, "uploads", "generated")
os.makedirs(GENERATED_DIR, exist_ok=True)

API_BASE_URL = os.getenv("API_BASE_URL", "http://127.0.0.1:8000")

# ---------------------------------------------------------------------------
# Font helper
# ---------------------------------------------------------------------------

_WINDOWS_FONTS = [
    r"C:\Windows\Fonts\calibri.ttf",
    r"C:\Windows\Fonts\arial.ttf",
    r"C:\Windows\Fonts\verdana.ttf",
]
_WINDOWS_SERIF_FONTS = [
    r"C:\Windows\Fonts\times.ttf",
    r"C:\Windows\Fonts\timesbd.ttf",
    r"C:\Windows\Fonts\georgia.ttf",
    r"C:\Windows\Fonts\cambria.ttc",
]
_LINUX_FONTS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
]
_LINUX_SERIF_FONTS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
]
_MAC_FONTS = [
    "/System/Library/Fonts/Helvetica.ttc",
    "/Library/Fonts/Arial.ttf",
]
_MAC_SERIF_FONTS = [
    "/System/Library/Fonts/Times.ttc",
    "/Library/Fonts/Georgia.ttf",
]

_DOC_NAME_PT = 52
_CERT_ID_PT = 18
_QR_MM = 60
_DOC_DPI = 96


def _pt_to_px(value: int, dpi: int = _DOC_DPI) -> int:
    return max(1, round(value * dpi / 72))


def _mm_to_px(value: int, dpi: int = _DOC_DPI) -> int:
    return max(1, round(value * dpi / 25.4))


def _get_font(size: int = 28) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in _WINDOWS_FONTS + _LINUX_FONTS + _MAC_FONTS:
        if os.path.exists(path):
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def _get_serif_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in _WINDOWS_SERIF_FONTS + _LINUX_SERIF_FONTS + _MAC_SERIF_FONTS:
        if os.path.exists(path):
            return ImageFont.truetype(path, size)
    return _get_font(size)


def _get_sans_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    return _get_font(size)


def _get_small_font() -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    return _get_font(16)


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def parse_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse an Excel file and return a list of row dicts.

    Required columns (case-insensitive): student_name, email, certificate_id.
    Missing certificate_id values are auto-generated.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    raw_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    # Normalise header names (lower-case, strip spaces)
    headers = [h.lower().replace(" ", "_") for h in raw_headers]

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue  # skip blank rows
        record: dict[str, Any] = dict(zip(headers, row))

        # Ensure required fields exist
        record.setdefault("student_name", "Unknown Student")
        record.setdefault("email", "")
        if not record.get("certificate_id"):
            record["certificate_id"] = str(uuid.uuid4())[:8].upper()

        rows.append(record)

    return rows


# ---------------------------------------------------------------------------
# Core generation
# ---------------------------------------------------------------------------

async def _download_image(url: str) -> bytes:
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(url)
        resp.raise_for_status()
        return resp.content


def _build_qr(verify_url: str, size: int = 120) -> Image.Image:
    qr = qrcode.QRCode(box_size=4, border=2)
    qr.add_data(verify_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    return img.resize((size, size))


def _draw_centered_text(
    draw: ImageDraw.ImageDraw,
    *,
    text: str,
    x: int,
    y: int,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    fill: tuple[int, int, int],
) -> None:
    try:
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        draw.text((x - text_width // 2, y - text_height // 2), text, fill=fill, font=font)
    except Exception:
        draw.text((x, y), text, fill=fill, font=font)


async def generate_certificates(
    template_url: str,
    layout_config: dict[str, Any],
    students: list[dict[str, Any]],
    verify_base_url: str = "http://localhost:5173/verify",
) -> list[dict[str, str]]:
    """Generate one PNG certificate per student.

    Returns:
        List of ``{"id": cert_id, "url": download_url, "student_name": name}``
    """

    # Download and open template
    img_bytes = await _download_image(template_url)
    try:
        bg_master = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
    except Exception as exc:
        raise ValueError(f"Could not open template image: {exc}") from exc

    width, height = bg_master.size

    # Resolve layout_config keys (support both nameX/nameY and placeholderX/Y)
    name_x_pct = float(layout_config.get("nameX", layout_config.get("placeholderX", 40)))
    name_y_pct = float(layout_config.get("nameY", layout_config.get("placeholderY", 36)))
    qr_x_pct = float(layout_config.get("qrX", 82))
    qr_y_pct = float(layout_config.get("qrY", 76))
    id_x_pct = float(layout_config.get("idX", 12))
    id_y_pct = float(layout_config.get("idY", 92))

    show_name = bool(layout_config.get("showStudentName", True))
    show_qr = bool(layout_config.get("showQR", True))
    show_id = bool(layout_config.get("showID", True))

    font_name = _get_serif_font(_pt_to_px(_DOC_NAME_PT))
    font_id = _get_sans_font(_pt_to_px(_CERT_ID_PT))

    results: list[dict[str, str]] = []

    for student in students:
        bg = bg_master.copy()
        draw = ImageDraw.Draw(bg)

        cert_id: str = str(student.get("certificate_id", str(uuid.uuid4())[:8].upper()))
        student_name: str = str(student.get("student_name", "Unknown"))

        # Convert percent to pixels (centre-anchored)
        name_x = int(width * name_x_pct / 100)
        name_y = int(height * name_y_pct / 100)
        qr_x = int(width * qr_x_pct / 100)
        qr_y = int(height * qr_y_pct / 100)
        id_x = int(width * id_x_pct / 100)
        id_y = int(height * id_y_pct / 100)

        # Draw student name (centre-anchored to match preview)
        if show_name:
            _draw_centered_text(
                draw,
                text=student_name,
                x=name_x,
                y=name_y,
                fill=(20, 20, 80),
                font=font_name,
            )

        # Paste QR code
        if show_qr:
            verify_url = f"{verify_base_url}/{cert_id}"
            qr_img = _build_qr(verify_url, size=_mm_to_px(_QR_MM))
            qr_half = qr_img.size[0] // 2
            qr_pos = (max(0, qr_x - qr_half), max(0, qr_y - qr_half))
            bg.paste(qr_img, qr_pos, mask=qr_img)

        # Draw certificate ID
        if show_id:
            id_text = f"ID: {cert_id}"
            _draw_centered_text(
                draw,
                text=id_text,
                x=id_x,
                y=id_y,
                fill=(56, 56, 56),
                font=font_id,
            )

        # Save as RGB PNG
        out_name = f"{cert_id}.png"
        out_path = os.path.join(GENERATED_DIR, out_name)
        bg.convert("RGB").save(out_path, "PNG")

        results.append(
            {
                "id": cert_id,
                "url": f"{API_BASE_URL}/uploads/generated/{out_name}",
                "student_name": student_name,
            }
        )

    return results
